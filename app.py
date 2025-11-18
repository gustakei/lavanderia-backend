# app.py — VERSÃO DEFINITIVA PARA SISTEMA SÃO GERALDO
from flask import Flask, request, jsonify, Response, stream_with_context
from flask_cors import CORS
import os
import json
import base64
import tempfile
from datetime import datetime, timedelta
from urllib.parse import urlparse, parse_qs
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from apscheduler.schedulers.background import BackgroundScheduler
import atexit
import logging

# ========================= CONFIGURAÇÃO =========================
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

# Configuração de disco
DATA_DIR = "/data"
try:
    os.makedirs(DATA_DIR, exist_ok=True)
    test_file = os.path.join(DATA_DIR, "test.txt")
    with open(test_file, 'w') as f:
        f.write("test")
    os.remove(test_file)
    logger.info(f"Disco configurado: {DATA_DIR}")
except Exception as e:
    DATA_DIR = tempfile.mkdtemp(prefix="lavanderia_")
    logger.warning(f"Falha no disco principal, usando: {DATA_DIR}")

CONFIG_FILE = os.path.join(DATA_DIR, "config.json")
HOSPITALS_FILE = os.path.join(DATA_DIR, "hospitals.json")
RELATORIO_PATH = os.path.join(DATA_DIR, "RelatorioLavanderiaSemanal.xlsx")

# Variáveis globais
config = {}
hospitals = []

# ========================= CARREGAR DADOS =========================
def load_data():
    global config, hospitals
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config.update(json.load(f))
                logger.info("Configurações carregadas")
        
        if os.path.exists(HOSPITALS_FILE):
            with open(HOSPITALS_FILE, 'r', encoding='utf-8') as f:
                hospitals.extend(json.load(f))
                logger.info(f"{len(hospitals)} hospitais carregados")
    except Exception as e:
        logger.error(f"Erro ao carregar dados: {e}")

load_data()

def save_data():
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        with open(HOSPITALS_FILE, 'w', encoding='utf-8') as f:
            json.dump(hospitals, f, indent=2, ensure_ascii=False)
        logger.info("Dados salvos com sucesso")
    except Exception as e:
        logger.error(f"Erro ao salvar dados: {e}")

# ========================= CÁLCULO SEMANA ANTERIOR =========================
def calcular_semana_anterior():
    hoje = datetime.now()
    dias_desde_segunda = hoje.weekday()
    inicio = hoje - timedelta(days=dias_desde_segunda + 7)
    fim = inicio + timedelta(days=6)
    return inicio, fim

# ========================= SESSÃO HTTP PERSISTENTE =========================
def criar_sessao():
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'pt-BR,pt;q=0.9,en;q=0.8',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1'
    })
    return session

def fazer_login(session):
    try:
        # URL de login do sistema São Geraldo
        login_url = "https://sistemasaogeraldoservice.com.br/sistema/Login.aspx"
        
        # Primeiro, acessar a página de login para obter os campos hidden
        response = session.get(login_url)
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Extrair campos hidden necessários para o ASP.NET
        viewstate = soup.find('input', {'name': '__VIEWSTATE'})
        viewstate = viewstate['value'] if viewstate else ''
        
        viewstategenerator = soup.find('input', {'name': '__VIEWSTATEGENERATOR'})
        viewstategenerator = viewstategenerator['value'] if viewstategenerator else ''
        
        eventvalidation = soup.find('input', {'name': '__EVENTVALIDATION'})
        eventvalidation = eventvalidation['value'] if eventvalidation else ''
        
        # Dados do formulário de login
        login_data = {
            '__VIEWSTATE': viewstate,
            '__VIEWSTATEGENERATOR': viewstategenerator,
            '__EVENTVALIDATION': eventvalidation,
            'txtUsuario': config.get('username', 'Guilherme Duarte'),
            'txtSenha': config.get('password', '13072006'),
            'Button1': 'Acessar'
        }
        
        # Fazer login
        response = session.post(login_url, data=login_data, allow_redirects=True)
        
        # Verificar se o login foi bem-sucedido
        # Se redirecionou para Default.aspx, provavelmente deu certo
        if 'Default.aspx' in response.url or response.status_code == 200:
            logger.info("Login realizado com sucesso")
            return True
        else:
            logger.error("Falha no login - possíveis credenciais incorretas")
            return False
            
    except Exception as e:
        logger.error(f"Erro durante o login: {e}")
        return False

# ========================= EXTRAÇÃO DE DADOS OTIMIZADA =========================
def extrair_dados_semana_anterior(session, hospital):
    try:
        inicio, fim = calcular_semana_anterior()
        periodo_text = f"{inicio.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}"

        # Extrair cliente_id da URL do hospital
        parsed = urlparse(hospital['url'])
        params = parse_qs(parsed.query)
        cliente_id = params.get('cliente', [None])[0]
        
        if not cliente_id:
            logger.warning(f"Cliente ID não encontrado para {hospital['name']}")
            return 0.0, [], periodo_text

        base_url = "https://sistemasaogeraldoservice.com.br/sistema/ListagemLavanderia.aspx"
        dados_por_mes = {}
        
        # Agrupar dias por mês/ano para otimizar as requisições
        current = inicio
        while current <= fim:
            mes_ano = current.strftime("%m/%Y")
            dia_str = current.strftime("%d/%m/%Y")
            dados_por_mes.setdefault(mes_ano, []).append(dia_str)
            current += timedelta(days=1)

        total_kg = 0.0
        todos_dados = []

        for mes_ano, dias in dados_por_mes.items():
            url = f"{base_url}?cliente={cliente_id}&periodo={mes_ano}"
            logger.info(f"Acessando: {url}")
            
            response = session.get(url)
            if response.status_code != 200:
                logger.warning(f"Falha ao acessar página para {mes_ano}")
                continue

            soup = BeautifulSoup(response.content, 'html.parser')
            
            # ENCONTRAR A TABELA CORRETA - ESTRATÉGIA ROBUSTA
            tabela = None
            
            # Tentar pelo ID primeiro
            tabela = soup.find('table', id='tabpedidos')
            
            # Se não encontrou, buscar por classe
            if not tabela:
                tabela = soup.find('table', class_='tabpedidos')
            
            # Se ainda não encontrou, buscar qualquer tabela que tenha a estrutura correta
            if not tabela:
                for table in soup.find_all('table'):
                    headers = table.find_all('th')
                    for header in headers:
                        if 'DIA' in header.get_text().upper():
                            tabela = table
                            break
                    if tabela:
                        break

            if not tabela:
                logger.warning(f"Tabela não encontrada para {hospital['name']} - {mes_ano}")
                continue

            # EXTRAIR DADOS DA TABELA
            for linha in tabela.find_all('tr')[1:]:  # Pular cabeçalho
                cols = linha.find_all('td')
                if len(cols) < 2: 
                    continue
                    
                data_str = cols[0].get_text(strip=True)
                
                # Verificar se esta data está no período que nos interessa
                if data_str not in dias: 
                    continue
                
                # Extrair valor de kg - segunda coluna (índice 1)
                kg_text = cols[1].get_text(strip=True)
                
                # Limpar e converter o texto para float
                kg_text = kg_text.replace('.', '').replace(' ', '').replace(',', '.')
                try: 
                    kg = float(kg_text)
                except ValueError: 
                    kg = 0.0
                    logger.warning(f"Valor inválido para {data_str}: {kg_text}")
                
                todos_dados.append({'data': data_str, 'kg': kg})
                total_kg += kg

        logger.info(f"Extraídos {total_kg:.2f} kg para {hospital['name']} - {periodo_text}")
        return total_kg, todos_dados, periodo_text
        
    except Exception as e:
        logger.error(f"Erro na extração para {hospital['name']}: {e}")
        return 0.0, [], periodo_text

# ========================= RELATÓRIO EXCEL =========================
def gerar_relatorio(resultados):
    try:
        # Criar workbook manualmente (sem pandas)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório Lavanderia"
        
        # Cabeçalhos
        ws.append(['Hospital', 'Período', 'Total (Kg)'])
        
        # Formatar cabeçalho
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Adicionar dados
        for resultado in resultados:
            ws.append([
                resultado['hospital'],
                resultado['periodo'], 
                resultado['total']
            ])
        
        # Total geral
        total_geral = sum(r['total'] for r in resultados)
        ws.append(['', 'Total Geral', total_geral])
        ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
        
        # Gráfico
        if len(resultados) > 1:
            chart = BarChart()
            data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row-1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row-1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.title = "Totais por Hospital"
            chart.style = 13
            ws.add_chart(chart, "E2")
        
        wb.save(RELATORIO_PATH)
        logger.info(f"Relatório gerado: {RELATORIO_PATH}")
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório: {e}")
        raise

# ========================= ENVIO DE EMAIL =========================
def enviar_email(destinatario):
    try:
        email = os.getenv('EMAIL_SEU')
        senha = os.getenv('EMAIL_SENHA')
        
        if not email or not senha:
            logger.warning("Credenciais de email não configuradas")
            return False
            
        if not destinatario:
            logger.warning("Destinatário não configurado")
            return False

        # Criar mensagem
        msg = MIMEMultipart()
        msg['From'] = email
        msg['To'] = destinatario
        msg['Subject'] = f"Relatório Lavanderia - {datetime.now().strftime('%d/%m/%Y')}"

        # Corpo do email
        corpo = f"""
        Prezado(a),
        
        Segue em anexo o relatório semanal da lavanderia.
        
        Período: {datetime.now().strftime('%d/%m/%Y')}
        
        Atenciosamente,
        Sistema Automatizado de Lavanderia
        """
        msg.attach(MIMEText(corpo, 'plain'))

        # Anexar arquivo
        with open(RELATORIO_PATH, "rb") as f:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="RelatorioLavanderia.xlsx"')
            msg.attach(part)

        # Enviar email
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(email, senha)
        server.sendmail(email, destinatario, msg.as_string())
        server.quit()
        
        logger.info("Email enviado com sucesso")
        return True
        
    except Exception as e:
        logger.error(f"Erro ao enviar email: {e}")
        return False

# ========================= AGENDAMENTO =========================
scheduler = BackgroundScheduler()
scheduler.start()
atexit.register(lambda: scheduler.shutdown(wait=False))

def executar_relatorio_agendado():
    if not hospitals:
        logger.warning("Agendamento: Nenhum hospital cadastrado")
        return
    
    if not config.get('email'):
        logger.warning("Agendamento: Email não configurado")
        return
    
    logger.info("Executando relatório agendado...")
    
    try:
        session = criar_sessao()
        if not fazer_login(session):
            logger.error("Agendamento: Falha no login")
            return

        resultados = []
        for hospital in hospitals:
            kg, dados, periodo = extrair_dados_semana_anterior(session, hospital)
            resultados.append({
                'hospital': hospital['name'], 
                'periodo': periodo, 
                'total': kg, 
                'dados': dados
            })

        gerar_relatorio(resultados)
        enviar_email(config['email'])
        logger.info("Relatório agendado executado com sucesso")
        
    except Exception as e:
        logger.error(f"Erro durante execução agendada: {e}")

def reagendar():
    try:
        scheduler.remove_all_jobs()
        horario = config.get('schedule', '').strip()
        
        if not horario:
            logger.info("Nenhum agendamento configurado")
            return

        if horario.startswith('cron['):
            try:
                params = horario[5:-1].strip()
                dia, hora = params.split(' ', 1)
                h, m = map(int, hora.split(':'))
                
                scheduler.add_job(
                    executar_relatorio_agendado,
                    'cron',
                    day_of_week=dia,
                    hour=h, 
                    minute=m,
                    id='recorrente'
                )
                logger.info(f"Agendamento recorrente configurado: {dia} às {h:02d}:{m:02d}")
                
            except Exception as e:
                logger.error(f"Erro ao configurar agendamento recorrente: {e}")
        else:
            try:
                dt = datetime.fromisoformat(horario.replace('Z', '+00:00'))
                if dt > datetime.now():
                    scheduler.add_job(executar_relatorio_agendado, 'date', run_date=dt)
                    logger.info(f"Agendamento único configurado para: {dt}")
            except Exception as e:
                logger.error(f"Erro ao configurar agendamento único: {e}")
                
    except Exception as e:
        logger.error(f"Erro no reagendamento: {e}")

# ========================= ROTAS =========================
@app.route('/')
def health_check():
    return jsonify({
        'status': 'online', 
        'data_dir': DATA_DIR,
        'hospitals_count': len(hospitals),
        'method': 'requests+beautifulsoup'
    })

@app.route('/api/data', methods=['GET'])
def get_data():
    return jsonify({'hospitals': hospitals, 'config': config})

@app.route('/api/config', methods=['POST'])
def update_config():
    global config
    try:
        config = request.json or {}
        save_data()
        reagendar()
        return jsonify({'status': 'ok', 'message': 'Configurações atualizadas'})
    except Exception as e:
        logger.error(f"Erro ao atualizar configurações: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/hospitals', methods=['POST'])
def add_hospital():
    global hospitals
    try:
        data = request.json
        if data and 'name' in data and 'url' in data:
            hospitals.append(data)
            save_data()
            return jsonify({'status': 'ok', 'message': 'Hospital adicionado'})
        return jsonify({'status': 'error', 'message': 'Dados inválidos'}), 400
    except Exception as e:
        logger.error(f"Erro ao adicionar hospital: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/hospitals/<int:index>', methods=['DELETE'])
def remove_hospital(index):
    global hospitals
    try:
        if 0 <= index < len(hospitals):
            removed = hospitals.pop(index)
            save_data()
            return jsonify({'status': 'ok', 'message': f'Hospital {removed["name"]} removido'})
        return jsonify({'status': 'error', 'message': 'Índice inválido'}), 404
    except Exception as e:
        logger.error(f"Erro ao remover hospital: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/run-stream', methods=['GET'])
def run_stream():
    def event(msg):
        return f"data: {json.dumps(msg, default=str)}\n\n"

    @stream_with_context
    def generate():
        if not hospitals:
            yield event({'type': 'error', 'error': 'Nenhum hospital cadastrado'})
            return

        yield event({'type': 'meta', 'total': len(hospitals)})

        try:
            session = criar_sessao()
            yield event({'type': 'progress', 'status': 'Iniciando sessão...'})
            
            yield event({'type': 'progress', 'status': 'Fazendo login no sistema...'})
            if not fazer_login(session):
                yield event({'type': 'error', 'error': 'Falha no login - verifique usuário e senha'})
                return

            resultados = []
            for idx, hospital in enumerate(hospitals, 1):
                yield event({
                    'type': 'progress', 
                    'idx': idx, 
                    'hospital': hospital['name'], 
                    'status': 'Conectando...'
                })
                
                kg, dados, periodo = extrair_dados_semana_anterior(session, hospital)
                resultados.append({
                    'hospital': hospital['name'], 
                    'periodo': periodo, 
                    'total': kg, 
                    'dados': dados
                })
                
                yield event({
                    'type': 'progress', 
                    'idx': idx, 
                    'hospital': hospital['name'], 
                    'status': f'{kg:.2f} kg'
                })

            yield event({'type': 'progress', 'status': 'Gerando relatório Excel...'})
            gerar_relatorio(resultados)

            if config.get('email'):
                yield event({'type': 'progress', 'status': 'Enviando email...'})
                enviar_email(config['email'])

            # Enviar Excel para frontend
            with open(RELATORIO_PATH, 'rb') as f:
                b64 = base64.b64encode(f.read()).decode()
                yield event({
                    'type': 'excel', 
                    'data': b64, 
                    'filename': 'RelatorioLavanderia.xlsx'
                })

            yield event({'type': 'done', 'results': resultados})
            
        except Exception as e:
            logger.error(f"Erro na execução: {e}")
            yield event({'type': 'error', 'error': str(e)})

    return Response(generate(), mimetype='text/event-stream')

# Rota para teste rápido do sistema
@app.route('/api/test-system', methods=['GET'])
def test_system():
    """Teste rápido para verificar se o sistema está funcionando"""
    try:
        session = criar_sessao()
        login_success = fazer_login(session)
        
        return jsonify({
            'status': 'success',
            'login': 'success' if login_success else 'failed',
            'data_dir': DATA_DIR,
            'hospitals_count': len(hospitals),
            'config_loaded': bool(config)
        })
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

if __name__ == '__main__':
    logger.info("Iniciando aplicação Flask...")
    reagendar()
    port = int(os.getenv('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)

